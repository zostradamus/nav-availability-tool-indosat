#!/usr/bin/env python3
"""
NAV 2G Availability Analysis - Central Java (IOH Region)
=========================================================
Pipeline otomatis:
1. Scan folder CommonCenter_NetworkAvailibility/Week_*_* untuk file .xlsb baru
2. Baca sheet 2G, filter New_Region = CENTRAL JAVA
3. Hitung availability rata-rata 7 hari per site
4. Tambah kolom Kelurahan (point-in-polygon dari koordinat site,
   batas kelurahan Jateng+DIY, sumber: github.com/cahyadsn/wilayah_boundaries)
5. Join atribut site dari CMDB (CommonDatabase_Gov, file minggu terbaru)
6. Kategori: 100% | 98-<100% | <98%  + Top 10 worst
7. Output Excel: NAV_2G_CentralJava_Analysis.xlsx

Dipakai oleh scheduled task - hanya file BARU yang diproses (cache per minggu).
Jalankan: python3 run_nav_analysis.py [--force] [--max-files N] [--no-excel]
"""
import os, sys, re, glob, json, pickle, subprocess, time, argparse

BASE = os.path.dirname(os.path.abspath(__file__))

_DEFAULT_CFG = {
    "nav_dir": "../CommonCenter_NetworkAvailibility",
    "cmdb_dir": "../CommonDatabase_Gov",
    "output_xlsx": "NAV_2G_CentralJava_Analysis.xlsx",
    "region": "CENTRAL JAVA",
    "output_dashboard": "dashboard.html",
    "copy_to": []
}
_cfg_path = os.path.join(BASE, "config.json")
cfg = dict(_DEFAULT_CFG)
if os.path.exists(_cfg_path):
    with open(_cfg_path) as _f:
        _raw = _f.read()
    try:
        cfg.update(json.loads(_raw))
    except json.JSONDecodeError:
        # toleransi path Windows dengan backslash tunggal (E:\folder\...)
        cfg.update(json.loads(_raw.replace("\\", "/")))

def _abs(p):
    return p if os.path.isabs(p) else os.path.normpath(os.path.join(BASE, p))

NAV_DIR = _abs(cfg["nav_dir"])
CMDB_DIR = _abs(cfg["cmdb_dir"])
OUT_XLSX = _abs(cfg["output_xlsx"])
REGION = cfg["region"]
DASH_OUT = _abs(cfg.get("output_dashboard", "dashboard.html"))
COPY_TO = [_abs(p) for p in cfg.get("copy_to", [])]
CACHE = os.path.join(BASE, "cache")
REF = os.path.join(BASE, "ref")
STATE_F = os.path.join(CACHE, "state.json")
KEL_CACHE_F = os.path.join(CACHE, "kelurahan_cache.json")
BOUND_PKL = os.path.join(REF, "kelurahan_boundaries_33_34.pkl")

def ensure_deps():
    for mod, pkg in [("pyxlsb", "pyxlsb"), ("shapely", "shapely"),
                     ("openpyxl", "openpyxl"), ("pandas", "pandas"),
                     ("pyarrow", "pyarrow")]:
        try:
            __import__(mod)
        except ImportError:
            r = subprocess.run([sys.executable, "-m", "pip", "install", pkg,
                                "--break-system-packages", "-q"])
            if r.returncode != 0:  # pip lama (mis. Windows) tak kenal flag itu
                subprocess.run([sys.executable, "-m", "pip", "install", pkg,
                                "-q"], check=True)

def load_json(path, default):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return default

def save_json(path, obj):
    with open(path, "w") as f:
        json.dump(obj, f)

def find_week_files():
    out = {}
    for d in sorted(glob.glob(os.path.join(NAV_DIR, "Week_*_*"))):
        m = re.match(r"Week_(\d+)_(\d+)", os.path.basename(d))
        if not m:
            continue
        wk = f"{m.group(2)}_W{int(m.group(1)):02d}"
        files = sorted(glob.glob(os.path.join(d, "*.xlsb")))
        if files:
            out[wk] = files[-1]
    return out

class KelurahanLookup:
    def __init__(self):
        from shapely import wkb
        from shapely.strtree import STRtree
        with open(BOUND_PKL, "rb") as f:
            rows = pickle.load(f)
        self.names = [r[1] for r in rows]
        self.kodes = [r[0] for r in rows]
        geoms = [wkb.loads(r[4]) for r in rows]
        self.tree = STRtree(geoms)
        self.geoms = geoms
        self.cache = load_json(KEL_CACHE_F, {})

    def lookup(self, siteid, lng, lat):
        key = f"{siteid}|{round(float(lng),5)}|{round(float(lat),5)}"
        if key in self.cache:
            return self.cache[key]
        from shapely.geometry import Point
        p = Point(float(lng), float(lat))
        idxs = self.tree.query(p, predicate="within")
        if len(idxs):
            i = int(idxs[0])
        else:
            i = int(self.tree.nearest(p))
            from shapely import distance
            if distance(self.geoms[i], p) > 0.05:  # >~5km dari kelurahan terdekat
                self.cache[key] = ["OUT_OF_AREA", ""]
                return self.cache[key]
        self.cache[key] = [self.names[i], self.kodes[i]]
        return self.cache[key]

    def save(self):
        save_json(KEL_CACHE_F, self.cache)

def process_week(week, path, kel):
    import pandas as pd
    need = {"SITEID", "sitenname", "MC Cluster", "New_Region", "Long", "Lat",
            "Kecamatan", "Kabupaten_Kota"}

    _xf = pd.ExcelFile(path, engine="pyxlsb")
    _names = _xf.sheet_names

    def _resolve(token):
        for n in _names:  # persis (abaikan kapital/spasi)
            if n.strip().upper() == token.upper():
                return n
        for n in _names:  # mengandung token, mis. "NAV 2G"
            if token.upper() in n.upper():
                return n
        raise ValueError(f"sheet '{token}' tidak ada di "
                         f"{os.path.basename(path)}; sheet tersedia: {_names}")

    def read_sheet(sheet, navpat, trafpat):
        sheet = _resolve(sheet)
        d = pd.read_excel(_xf, sheet_name=sheet,
                          usecols=lambda c: c in need or
                          (isinstance(c, str) and
                           (re.match(navpat, c) or re.match(trafpat, c))))
        d = d[d["New_Region"] == REGION].copy()
        d = d.drop_duplicates("SITEID")
        navc = sorted(c for c in d.columns if re.match(navpat, c))
        trafc = sorted(c for c in d.columns if re.match(trafpat, c))
        for c in navc + trafc:
            d[c] = pd.to_numeric(d[c], errors="coerce")
        return d, navc, trafc

    g2, navc2, trafc2 = read_sheet("2G", r"^2G_NAV_\d{8}$",
                                   r"^IOH_TRAFFIC_DATA_MBYTE_\d{8}$")
    g4, navc4, trafc4 = read_sheet("4G", r"^4G_NAV_\d{8}$",
                                   r"^IOH_DATA_TRAFFIC_MBYTE_4G_\d{8}$")

    g2["Avail_Avg"] = g2[navc2].mean(axis=1, skipna=True)  # NaN = tanpa data NAV 2G
    g2["Traffic_2G_MB"] = g2[trafc2].sum(axis=1, min_count=1)
    g2["In_2G_Sheet"] = True
    g4a = pd.DataFrame({
        "SITEID": g4["SITEID"],
        "Avail_4G": g4[navc4].mean(axis=1, skipna=True),
        "Traffic_4G_MB": g4[trafc4].sum(axis=1, min_count=1)})

    df = g2.drop(columns=trafc2).merge(g4a, on="SITEID", how="outer")
    df["In_2G_Sheet"] = df["In_2G_Sheet"].fillna(False)
    meta4 = g4.drop_duplicates("SITEID").set_index("SITEID")
    only4 = df["sitenname"].isna()
    for c in ["sitenname", "MC Cluster", "Long", "Lat", "Kecamatan",
              "Kabupaten_Kota"]:
        df.loc[only4, c] = df.loc[only4, "SITEID"].map(meta4[c])

    kelu, kode = [], []
    for _, r in df.iterrows():
        if pd.isna(r["Long"]) or pd.isna(r["Lat"]):
            kelu.append("NO_COORD"); kode.append("")
        else:
            n, k = kel.lookup(r["SITEID"], r["Long"], r["Lat"])
            kelu.append(n); kode.append(k)
    df["Kelurahan"] = kelu
    df["Kode_Kelurahan"] = kode
    ren = {c: "NAV_" + c[-8:] for c in navc2}
    df = df.rename(columns=ren)
    cols = ["SITEID", "sitenname", "MC Cluster", "Kelurahan", "Kode_Kelurahan",
            "Kecamatan", "Kabupaten_Kota", "Long", "Lat"] + \
           [ren[c] for c in navc2] + \
           ["Avail_Avg", "Avail_4G", "Traffic_2G_MB", "Traffic_4G_MB",
            "In_2G_Sheet"]
    df = df[cols].rename(columns={"sitenname": "Site_Name",
                                  "Kabupaten_Kota": "Kabupaten"})
    df.insert(0, "Week", week)
    df.insert(1, "Week_Date", min(c[-8:] for c in navc2))
    return df

CMDB_COLS = {"Site ID*": "SITEID_CM", "Site Type Label": "Site_Type",
             "Transport Type Label": "Transport_Type",
             "Site Priority Label": "Site_Priority", "Is VIP": "Is_VIP",
             "B2B Site": "B2B_Site", "Has Generator": "Has_Generator",
             "Battery Backup Category": "Battery_Backup", "Owner": "Owner",
             "Site Address": "Site_Address"}

def find_cmdb_file():
    best = None
    for f in glob.glob(os.path.join(CMDB_DIR, "*.xlsx")):
        name = os.path.basename(f)
        m = re.search(r"W(\d+)", name)
        if m:
            mv = re.search(r"\bv(\d+)\b", name, re.I)  # "W20 v2" > "W20"
            key = (int(m.group(1)), int(mv.group(1)) if mv else 1, name)
            if best is None or key > best[0]:
                best = (key, f)
    return best[1] if best else None

def process_cmdb(path):
    import pandas as pd
    cm = pd.read_excel(path, usecols=list(CMDB_COLS)).rename(columns=CMDB_COLS)
    cm["SITEID_CM"] = cm["SITEID_CM"].astype(str).str.strip().str.upper()
    cm = cm.drop_duplicates("SITEID_CM")
    cm.to_parquet(os.path.join(CACHE, "cmdb.parquet"), index=False)
    return cm

def merge_cmdb(df):
    import pandas as pd
    p = os.path.join(CACHE, "cmdb.parquet")
    if not os.path.exists(p):
        return df.copy()
    cm = pd.read_parquet(p)
    df = df.copy()
    df["_sid"] = df["SITEID"].astype(str).str.strip().str.upper()
    df = df.merge(cm, left_on="_sid", right_on="SITEID_CM", how="left")
    cm_cols = [c for c in cm.columns if c != "SITEID_CM"]
    df.loc[df["SITEID_CM"].isna(), cm_cols] = "NOT_IN_CMDB"
    return df.drop(columns=["_sid", "SITEID_CM"])

def categorize(v):
    if v >= 1.0 - 1e-9:
        return "1. Availability 100%"
    if v >= 0.98:
        return "2. Availability 98% - <100%"
    return "3. Availability <98%"

def build_excel():
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    files = sorted(glob.glob(os.path.join(CACHE, "week_*.parquet")))
    if not files:
        print("No cached weeks; nothing to build")
        return
    weeks = {}
    for f in files:
        wdf = pd.read_parquet(f)
        weeks[wdf["Week"].iloc[0]] = wdf
    wkeys = sorted(weeks)
    latest = wkeys[-1]
    df_all = merge_cmdb(weeks[latest])
    _cmdb_f = [c for c in ["Site_Type", "Transport_Type", "Site_Priority",
               "Is_VIP", "B2B_Site", "Has_Generator", "Battery_Backup",
               "Owner", "Site_Address"] if c in df_all.columns]
    for _c in _cmdb_f:  # kosong di master CMDB (bukan gagal join)
        df_all[_c] = df_all[_c].fillna("BLANK_IN_CMDB").replace("", "BLANK_IN_CMDB")
    no_nav = df_all[df_all["Avail_Avg"].isna() & df_all.get("In_2G_Sheet", True)].copy()
    df = df_all[df_all["Avail_Avg"].notna()].copy()
    df["Category"] = df["Avail_Avg"].apply(categorize)
    df["Avail_%"] = df["Avail_Avg"] * 100
    df = df.sort_values("Avail_Avg")

    HDR_FILL = PatternFill("solid", start_color="FFCC0000")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
    TXT = Font(name="Arial", size=10)
    thin = Side(style="thin", color="FFBBBBBB")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CAT_FILL = {"1. Availability 100%": "FFC6EFCE",
                "2. Availability 98% - <100%": "FFFFEB9C",
                "3. Availability <98%": "FFFFC7CE"}

    wb = Workbook()

    def write_df(ws, data, pct_cols=(), freeze="A2"):
        ws.append(list(data.columns))
        for c in ws[1]:
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        for row in data.itertuples(index=False):
            ws.append(list(row))
        ncol = len(data.columns)
        for j, colname in enumerate(data.columns, 1):
            L = get_column_letter(j)
            try:
                w = int(data[colname].astype(str).str.len().max()) + 2
            except Exception:
                w = 12
            ws.column_dimensions[L].width = max(10, min(30, w,
                                                        ), len(str(colname)) + 2)
            for i in range(2, len(data) + 2):
                cell = ws.cell(row=i, column=j)
                cell.font = TXT
                cell.border = BORDER
                if colname in pct_cols:
                    cell.number_format = "0.00%"
        ws.freeze_panes = freeze
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{len(data)+1}"

    nav_day_cols = [c for c in df.columns if c.startswith("NAV_")]
    cmdb_extra = [c for c in ["Site_Type", "Transport_Type", "Site_Priority",
                              "Is_VIP", "B2B_Site", "Has_Generator",
                              "Battery_Backup", "Owner", "Site_Address"]
                  if c in df.columns]
    detail_cols = ["Week", "SITEID", "Site_Name", "Kelurahan", "Kecamatan",
                   "Kabupaten", "MC Cluster"] + cmdb_extra + \
                  ["Long", "Lat"] + nav_day_cols + \
                  ["Avail_Avg", "Avail_4G", "Traffic_2G_MB", "Traffic_4G_MB",
                   "Category"]

    # ---- Sheet Data (semua site minggu terbaru, AVERAGE & kategori via formula)
    ws = wb.active
    ws.title = "Data_" + latest
    ddf = df[detail_cols].copy()
    write_df(ws, ddf, pct_cols=nav_day_cols + ["Avail_Avg"])
    first_nav = get_column_letter(detail_cols.index(nav_day_cols[0]) + 1)
    last_nav = get_column_letter(detail_cols.index(nav_day_cols[-1]) + 1)
    avg_col = detail_cols.index("Avail_Avg") + 1
    cat_col = detail_cols.index("Category") + 1
    for i in range(2, len(ddf) + 2):
        ws.cell(row=i, column=avg_col).value = \
            f"=AVERAGE({first_nav}{i}:{last_nav}{i})"
        a = f"{get_column_letter(avg_col)}{i}"
        ws.cell(row=i, column=cat_col).value = \
            (f'=IF({a}>=1,"1. Availability 100%",'
             f'IF({a}>=0.98,"2. Availability 98% - <100%",'
             f'"3. Availability <98%"))')
        ws.cell(row=i, column=cat_col).fill = \
            PatternFill("solid", start_color=CAT_FILL[df.iloc[i-2]["Category"]])
    GREY = PatternFill("solid", start_color="FFD9D9D9")
    RED = PatternFill("solid", start_color="FFFFC7CE")
    YEL = PatternFill("solid", start_color="FFFFEB9C")
    nav_idx = [detail_cols.index(c) + 1 for c in nav_day_cols]
    for i in range(2, len(ddf) + 2):
        rowv = df.iloc[i-2]
        for c, j in zip(nav_day_cols, nav_idx):
            v = rowv[c]
            if pd.isna(v):
                ws.cell(row=i, column=j).fill = GREY
            elif v < 0.98:
                ws.cell(row=i, column=j).fill = RED
            elif v < 1.0 - 1e-9:
                ws.cell(row=i, column=j).fill = YEL

    # ---- Summary
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "NAV 2G AVAILABILITY - CENTRAL JAVA (IOH REGION)"
    ws["B2"].font = Font(name="Arial", bold=True, size=14)
    ws["B3"] = f"Minggu terbaru: {latest}  |  Rata-rata 7 hari per site  |  " \
               f"Generated: {time.strftime('%Y-%m-%d %H:%M')}"
    ws["B3"].font = Font(name="Arial", italic=True, size=9, color="FF666666")
    ws["B5"] = "Kategori"; ws["C5"] = "Jumlah Site"; ws["D5"] = "% dari Total"
    for c in ("B5", "C5", "D5"):
        ws[c].fill = HDR_FILL; ws[c].font = HDR_FONT; ws[c].border = BORDER
    dname = "Data_" + latest
    catL = get_column_letter(cat_col)
    cats = ["1. Availability 100%", "2. Availability 98% - <100%",
            "3. Availability <98%"]
    for k, cat in enumerate(cats):
        r = 6 + k
        ws[f"B{r}"] = cat
        ws[f"C{r}"] = f"=COUNTIF('{dname}'!{catL}:{catL},B{r})"
        ws[f"D{r}"] = f"=C{r}/C$9"
        ws[f"D{r}"].number_format = "0.0%"
        for col in "BCD":
            ws[f"{col}{r}"].border = BORDER
            ws[f"{col}{r}"].font = TXT
        ws[f"B{r}"].fill = PatternFill("solid", start_color=CAT_FILL[cat])
    ws["B9"] = "Total site"
    ws["C9"] = "=SUM(C6:C8)"
    ws["B9"].font = Font(name="Arial", bold=True)
    ws["C9"].font = Font(name="Arial", bold=True)
    for col in "BCD":
        ws[f"{col}9"].border = BORDER
    ws["B11"] = "Rata-rata availability regional"
    ws["C11"] = f"=AVERAGE('{dname}'!{get_column_letter(avg_col)}2:" \
                f"{get_column_letter(avg_col)}{len(ddf)+1})"
    ws["C11"].number_format = "0.00%"
    ws["B11"].font = TXT; ws["C11"].font = TXT
    ws["B13"] = "KETERANGAN WARNA & KODE"
    ws["B13"].font = Font(name="Arial", bold=True, size=11)
    _legend = [
        ("FFC6EFCE", "Hijau", "Availability 100% (kategori 1)"),
        ("FFFFEB9C", "Kuning", "98% - <100% (kategori 2 / nilai harian di sheet Data)"),
        ("FFFFC7CE", "Merah", "<98% (kategori 3 / nilai harian di sheet Data)"),
        ("FFD9D9D9", "Abu-abu", "Sel NAV harian kosong = tidak ada data monitoring hari itu"),
    ]
    for k, (clr, nm, desc) in enumerate(_legend):
        r = 14 + k
        ws[f"B{r}"] = nm
        ws[f"B{r}"].fill = PatternFill("solid", start_color=clr)
        ws[f"B{r}"].font = TXT
        ws[f"B{r}"].border = BORDER
        ws[f"C{r}"] = desc
        ws[f"C{r}"].font = TXT
    _kode = [
        ("NOT_IN_CMDB", "Site tidak ditemukan di master CMDB (kemungkinan site baru / beda ID)"),
        ("BLANK_IN_CMDB", "Field ini memang belum diisi di master CMDB oleh tim data"),
        ("OUT_OF_AREA", "Koordinat site di luar Jateng+DIY - kemungkinan salah input koordinat"),
        ("NO_COORD", "Site tanpa koordinat Long/Lat di file NAV"),
        ("Sheet No_NAV_Data", "Site tanpa data NAV 7 hari penuh - tidak dihitung dalam kategori, perlu dicek tim monitoring"),
    ]
    for k, (kd, desc) in enumerate(_kode):
        r = 19 + k
        ws[f"B{r}"] = kd
        ws[f"B{r}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"B{r}"].border = BORDER
        ws[f"C{r}"] = desc
        ws[f"C{r}"].font = TXT
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12

    # ---- 3 sheet kategori
    sheet_map = [("Avail_100", cats[0]), ("Avail_98_99", cats[1]),
                 ("Below_98", cats[2])]
    cat_cols = ["SITEID", "Site_Name", "Kelurahan", "Kecamatan", "Kabupaten",
                "MC Cluster"] + [c for c in ["Transport_Type", "Site_Priority",
                "Has_Generator", "Owner"] if c in df.columns] + ["Avail_%"]
    for sname, cat in sheet_map:
        ws = wb.create_sheet(sname)
        sub = df[df["Category"] == cat][cat_cols].copy()
        sub["Avail_%"] = sub["Avail_%"].round(2)
        write_df(ws, sub)

    # ---- Top 10 worst
    ws = wb.create_sheet("Top10_Worst")
    top10_cols = ["SITEID", "Site_Name", "Kelurahan", "Kecamatan", "Kabupaten",
                  "MC Cluster"] + [c for c in ["Transport_Type",
                  "Site_Priority", "Has_Generator", "Owner"]
                  if c in df.columns] + ["Avail_%", "Category"]
    top = df.nsmallest(10, "Avail_Avg")[top10_cols].copy()
    top["Avail_%"] = top["Avail_%"].round(2)
    top.insert(0, "Rank", range(1, len(top) + 1))
    write_df(ws, top)
    cat_j = len(top.columns)
    for i in range(2, len(top) + 2):
        ws.cell(row=i, column=cat_j).fill = PatternFill(
            "solid", start_color=CAT_FILL[top.iloc[i-2]["Category"]])

    # ---- Site tanpa data NAV (7 hari kosong) - perlu dicek tim monitoring
    ws = wb.create_sheet("No_NAV_Data")
    nn_cols = [c for c in ["SITEID", "Site_Name", "Kelurahan", "Kecamatan",
               "Kabupaten", "MC Cluster", "Site_Type", "Transport_Type",
               "Site_Priority", "Owner"] if c in no_nav.columns]
    write_df(ws, no_nav[nn_cols].sort_values("Kabupaten"))

    # ---- Weekly trend (semua minggu di cache)
    ws = wb.create_sheet("Weekly_Trend")
    rows = []
    for wk in wkeys:
        w0 = weeks[wk]
        w = w0[w0["Avail_Avg"].notna()]
        cat = w["Avail_Avg"].apply(categorize)
        rows.append({"Week": wk, "Total_Site": len(w),
                     "No_NAV_Data": int((w0["Avail_Avg"].isna() & w0.get("In_2G_Sheet", True)).sum()),
                     "Avail_100": int((cat == cats[0]).sum()),
                     "Avail_98_99": int((cat == cats[1]).sum()),
                     "Below_98": int((cat == cats[2]).sum()),
                     "Avg_Avail_%": round(w["Avail_Avg"].mean() * 100, 2)})
    write_df(ws, pd.DataFrame(rows))

    out = OUT_XLSX
    try:
        wb.save(out)
    except PermissionError:
        out = OUT_XLSX.replace(".xlsx", "_new.xlsx")
        wb.save(out)
        print("WARNING: file utama terkunci (terbuka di Excel?), "
              "disimpan sebagai *_new.xlsx")
    print(f"Excel saved: {out} | latest={latest} | weeks={len(wkeys)} "
          f"| sites={len(df)}")
    import shutil
    for dest in COPY_TO:
        try:
            os.makedirs(dest, exist_ok=True)
            shutil.copy2(out, os.path.join(dest, os.path.basename(OUT_XLSX)))
            print(f"copied to: {dest}")
        except Exception as e:
            print(f"WARNING: gagal copy ke {dest}: {e}")

def build_dashboard():
    import pandas as pd
    tpl_p = os.path.join(BASE, "dashboard_template.html")
    if not os.path.exists(tpl_p):
        print("WARNING: dashboard_template.html tidak ada, dashboard dilewati")
        return
    files = sorted(glob.glob(os.path.join(CACHE, "week_*.parquet")))
    if not files:
        return
    weeks = {}
    for f in files:
        w = pd.read_parquet(f)
        weeks[w["Week"].iloc[0]] = w
    wkeys = sorted(weeks)
    wdates = [str(weeks[k]["Week_Date"].iloc[0]) if "Week_Date" in weeks[k]
              else "20260101" for k in wkeys]

    # registry site: meta dari kemunculan terbaru
    meta = {}
    for k in wkeys:
        w = weeks[k]
        for t in w.itertuples(index=False):
            meta[t.SITEID] = t
    sids = sorted(meta)
    idx = {s: j for j, s in enumerate(sids)}

    cmdb = {}
    p = os.path.join(CACHE, "cmdb.parquet")
    if os.path.exists(p):
        cm = pd.read_parquet(p)
        for t in cm.itertuples(index=False):
            cmdb[t.SITEID_CM] = t

    def _s(v):
        return "" if v is None or (isinstance(v, float)) or pd.isna(v) else str(v)

    sites = []
    for s in sids:
        t = meta[s]
        c = cmdb.get(str(s).strip().upper())
        sites.append([s, _s(t.Site_Name), _s(t.Kelurahan), _s(t.Kecamatan),
                      _s(t.Kabupaten), _s(t[4]),
                      _s(c.Site_Type) if c else "", _s(c.Transport_Type) if c else "",
                      _s(c.Site_Priority) if c else "", _s(c.Owner) if c else "",
                      _s(c.Has_Generator) if c else ""])

    def _r(v, nd=6):
        return None if pd.isna(v) else round(float(v), nd)

    data = {}
    for k in wkeys:
        arr = [None] * len(sids)
        w = weeks[k]
        a2 = w["Avail_Avg"].values; a4 = w["Avail_4G"].values
        t2 = w["Traffic_2G_MB"].values; t4 = w["Traffic_4G_MB"].values
        sid = w["SITEID"].values
        for j in range(len(w)):
            arr[idx[sid[j]]] = [_r(a2[j]), _r(a4[j]), _r(t2[j], 1), _r(t4[j], 1)]
        for j in range(len(arr)):
            if arr[j] is None:
                arr[j] = [None, None, None, None]
        data[k] = arr

    payload = {"generated": time.strftime("%Y-%m-%d %H:%M"), "region": REGION,
               "weeks": wkeys, "week_dates": wdates, "sites": sites,
               "data": data}
    tpl = open(tpl_p, encoding="utf-8").read()
    html = tpl.replace("__PAYLOAD__", json.dumps(payload, separators=(",", ":"),
                                                 ensure_ascii=False))
    out = DASH_OUT
    try:
        with open(out, "w", encoding="utf-8") as f:
            f.write(html)
    except PermissionError:
        out = DASH_OUT.replace(".html", "_new.html")
        with open(out, "w", encoding="utf-8") as f:
            f.write(html)
    print(f"Dashboard saved: {out} ({len(html)//1024} KB)")
    import shutil
    for dest in COPY_TO:
        try:
            os.makedirs(dest, exist_ok=True)
            shutil.copy2(out, os.path.join(dest, os.path.basename(DASH_OUT)))
        except Exception as e:
            print(f"WARNING: gagal copy dashboard ke {dest}: {e}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--max-files", type=int, default=99)
    ap.add_argument("--no-excel", action="store_true")
    ap.add_argument("--excel-only", action="store_true")
    args = ap.parse_args()

    ensure_deps()
    os.makedirs(CACHE, exist_ok=True)

    if not args.excel_only:
        state = load_json(STATE_F, {})
        cmdb_f = find_cmdb_file()
        if cmdb_f:
            sig = f"{os.path.basename(cmdb_f)}_{os.path.getmtime(cmdb_f)}"
            if args.force or state.get("__cmdb__") != sig:
                cm = process_cmdb(cmdb_f)
                state["__cmdb__"] = sig
                save_json(STATE_F, state)
                print(f"processed CMDB: {os.path.basename(cmdb_f)} "
                      f"({len(cm)} sites)")
        week_files = find_week_files()
        todo = []
        for wk, path in week_files.items():
            sig = f"{os.path.getmtime(path)}_{os.path.getsize(path)}"
            if args.force or state.get(wk) != sig:
                todo.append((wk, path, sig))
        todo = todo[:args.max_files]
        if todo:
            kel = KelurahanLookup()
            for wk, path, sig in todo:
                t0 = time.time()
                try:
                    df = process_week(wk, path, kel)
                except Exception as e:
                    print(f"ERROR {wk} ({os.path.basename(path)}): {e} "
                          f"-- minggu ini dilewati")
                    continue
                df.to_parquet(os.path.join(CACHE, f"week_{wk}.parquet"),
                              index=False)
                state[wk] = sig
                save_json(STATE_F, state)
                kel.save()
                print(f"processed {wk}: {len(df)} sites "
                      f"({time.time()-t0:.0f}s)")
        else:
            print("no new/changed week files")

    if not args.no_excel:
        build_excel()
        build_dashboard()

if __name__ == "__main__":
    main()
